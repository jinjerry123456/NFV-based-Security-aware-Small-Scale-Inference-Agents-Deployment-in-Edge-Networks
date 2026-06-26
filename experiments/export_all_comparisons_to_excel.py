# -*- coding: utf-8 -*-
"""
将六个实验脚本的对比结果导出到一个 Excel 文件中：

1. compare_algorithms_by_nodes.py        （自变量：节点个数）
2. compare_algorithms_by_degree.py       （自变量：拓扑度数）
3. compare_algorithms_by_clients.py      （自变量：客户端数量）
4. compare_algorithms_by_nf_instances.py （自变量：节点 NF instance 个数）
5. compare_algorithms_by_security.py     （自变量：安全需求）
6. compare_algorithms_by_client_acceptance.py （自变量：客户端数量，指标：接受率）

每个实验在 Excel 中对应一个工作表，内容等价于各自脚本在终端打印的摘要：
- 行：3 个指标 × (最小值/最大值/平均值)（接受率实验只有1个指标）
- 列：4 个算法（My / Cost-first / Price-first / Security-first）
"""

import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# 导入六个实验模块
import experiments.compare_algorithms_by_nodes as exp_nodes
import experiments.compare_algorithms_by_degree as exp_degree
import experiments.compare_algorithms_by_clients as exp_clients
import experiments.compare_algorithms_by_nf_instances as exp_nf
import experiments.compare_algorithms_by_security as exp_security
import experiments.compare_algorithms_by_client_acceptance as exp_client_acceptance


ALGORITHM_NAMES = ["My Algorithm", "Cost-first", "Price-first", "Security-first"]

# 三个统一的指标定义（展示名称, 内部ID, Excel 格式）+ 新增指标
METRICS = [
    ("总体成本/安全级别", "overall_cost_per_security", "{:.4f}"),
    ("平均安全级别", "avg_security", "{:.2f}"),
    ("总体成本 ($)", "overall_cost", "{:.2f}"),
    ("总体带宽成本 ($)", "overall_bandwidth_cost", "{:.2f}"),
    ("平均带宽成本 ($)", "avg_bandwidth_cost", "{:.4f}"),
    ("Client到Node平均跳数", "avg_client_node_hops", "{:.2f}"),
    ("总体安全补足成本 ($)", "overall_security_augmentation_cost", "{:.2f}"),
    ("总体扩容成本 ($)", "overall_augmentation_cost", "{:.2f}"),
    ("平均安全补足开销 ($)", "avg_security_augmentation_cost", "{:.4f}"),
    ("Node平均部署NF个数(补足)", "avg_augmented_nf_per_node", "{:.2f}"),
    ("补足SecurityLevel/补足Cost", "security_level_per_aug_cost", "{:.4f}"),
    ("补足NF个数/补足Cost", "nf_count_per_aug_cost", "{:.4f}"),
    ("Node平均扩容GPU unit", "avg_augmentation_gpu_per_node", "{:.4f}"),
    ("Node个数/扩容开销", "nodes_per_aug_cost", "{:.4f}"),
    ("Node平均Client覆盖率", "avg_clients_per_node", "{:.2f}"),
    ("算法执行时间 (s)", "execution_time", "{:.4f}"),
    ("使用的节点数", "total_nodes", "{:.0f}"),
    ("单节点平均成本 ($)", "avg_cost_per_node", "{:.4f}"),
]

# 指标分组（用于表格分区展示，便于阅读）
METRIC_SECTIONS = [
    ("成本与总体", ["overall_cost_per_security", "avg_security", "overall_cost", "avg_cost_per_node"]),
    ("带宽与路径", ["overall_bandwidth_cost", "avg_bandwidth_cost", "avg_client_node_hops"]),
    ("安全补足", ["overall_security_augmentation_cost", "overall_augmentation_cost", "avg_security_augmentation_cost", "avg_augmented_nf_per_node", "security_level_per_aug_cost", "nf_count_per_aug_cost"]),
    ("扩容与节点", ["avg_augmentation_gpu_per_node", "nodes_per_aug_cost", "total_nodes", "avg_clients_per_node"]),
    ("性能", ["execution_time"]),
]

# 由 metric_id 查 (展示名, 格式)
_metrics_by_id = {m[1]: (m[0], m[2]) for m in METRICS}

# 默认运行次数（可通过 main 参数或命令行覆盖）
DEFAULT_NUM_RUNS = 10
# 默认每个 size/degree 生成的随机拓扑数量（统一覆盖四个实验）
DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO = 3
# 默认是否打印详细进度信息（可通过 main 参数或命令行覆盖）
DEFAULT_VERBOSE_PROGRESS = True


def _write_metrics_block(ws, row_idx, dim_value, per_alg_metrics, header_fill, section_fill):
    """
    写入一个维度块（如「20 节点」下的所有指标），按分组展示，指标名合并三行，统计列单独一列。
    返回写入后的下一行号。
    """
    n_cols = 3 + len(ALGORITHM_NAMES)  # 维度 | 指标 | 统计 | 4 个算法
    section_font = Font(bold=True, size=10)
    for section_name, metric_ids in METRIC_SECTIONS:
        # 分组标题行
        ws.cell(row=row_idx, column=1, value=dim_value)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=2, value=f"【{section_name}】")
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        # 该分组下每个指标：3 行（最小/最大/平均），指标名合并
        for metric_id in metric_ids:
            if metric_id not in _metrics_by_id:
                continue
            metric_name, fmt = _metrics_by_id[metric_id]
            start_r = row_idx
            for stat_name, stat_index in (("最小值", 0), ("最大值", 1), ("平均值", 2)):
                ws.cell(row=row_idx, column=1, value=dim_value)
                ws.cell(row=row_idx, column=3, value=stat_name)
                for ci, alg_name in enumerate(ALGORITHM_NAMES):
                    mn, mx, avg = per_alg_metrics[alg_name].get(metric_id, (None, None, None))
                    val = (mn, mx, avg)[stat_index]
                    cell = ws.cell(row=row_idx, column=4 + ci, value=fmt.format(val) if isinstance(val, (int, float)) else "N/A")
                    cell.alignment = Alignment(horizontal="right")
                row_idx += 1
            ws.merge_cells(start_row=start_r, start_column=2, end_row=row_idx - 1, end_column=2)
            cell = ws.cell(row=start_r, column=2, value=metric_name)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    return row_idx


def _set_all_num_runs(num_runs: int):
    """
    将六个实验脚本中的 NUM_RUNS 统一设置为指定值。
    如果某个脚本没有 NUM_RUNS，则跳过。
    """
    for mod in (exp_nodes, exp_degree, exp_clients, exp_nf, exp_security, exp_client_acceptance):
        if hasattr(mod, "NUM_RUNS"):
            setattr(mod, "NUM_RUNS", int(num_runs))


def _set_all_generated_topologies_per_combo(count: int):
    """
    将六个实验脚本中的 GENERATED_TOPOLOGIES_PER_COMBO 统一设置为指定值。
    如果某个脚本没有该变量，则跳过。
    """
    for mod in (exp_nodes, exp_degree, exp_clients, exp_nf, exp_security, exp_client_acceptance):
        if hasattr(mod, "GENERATED_TOPOLOGIES_PER_COMBO"):
            setattr(mod, "GENERATED_TOPOLOGIES_PER_COMBO", int(count))


def _set_all_verbose_progress(verbose: bool):
    """
    将六个实验脚本中的 VERBOSE_PROGRESS 统一设置为指定值。
    如果某个脚本没有该变量，则跳过。
    """
    for mod in (exp_nodes, exp_degree, exp_clients, exp_nf, exp_security, exp_client_acceptance):
        if hasattr(mod, "VERBOSE_PROGRESS"):
            setattr(mod, "VERBOSE_PROGRESS", bool(verbose))


def _extract_metric_from_result(metric_id, res):
    """从单次算法返回的结果字典中计算一个指标值。"""
    if not isinstance(res, dict):
        return None
    try:
        if metric_id == "overall_cost_per_security":
            total_cost = res.get("total_cost")
            avg_sec = res.get("avg_node_security_level")
            if isinstance(total_cost, (int, float)) and isinstance(avg_sec, (int, float)) and avg_sec > 0:
                return total_cost / avg_sec
        elif metric_id == "avg_security":
            return res.get("avg_node_security_level")
        elif metric_id == "overall_cost":
            return res.get("total_cost")
        elif metric_id == "overall_bandwidth_cost":
            return res.get("total_bandwidth_cost")
        elif metric_id == "avg_bandwidth_cost":
            total_bw = res.get("total_bandwidth_cost")
            total_clients = res.get("total_clients")
            if isinstance(total_bw, (int, float)) and isinstance(total_clients, (int, float)) and total_clients > 0:
                return total_bw / total_clients
        elif metric_id == "avg_client_node_hops":
            return res.get("avg_client_node_hops")
        elif metric_id == "overall_security_augmentation_cost":
            return res.get("security_augmentation_cost")
        elif metric_id == "overall_augmentation_cost":
            return res.get("security_augmentation_cost")
        elif metric_id == "avg_security_augmentation_cost":
            aug_cost = res.get("security_augmentation_cost")
            total_clients = res.get("total_clients")
            if isinstance(aug_cost, (int, float)) and isinstance(total_clients, (int, float)) and total_clients > 0:
                return aug_cost / total_clients
        elif metric_id == "avg_augmented_nf_per_node":
            nf_count = res.get("total_augmented_nf_count")
            total_nodes = res.get("total_nodes")
            if isinstance(nf_count, (int, float)) and isinstance(total_nodes, (int, float)) and total_nodes > 0:
                return nf_count / total_nodes
        elif metric_id == "security_level_per_aug_cost":
            gain = res.get("total_security_level_gain")
            aug_cost = res.get("security_augmentation_cost")
            if isinstance(gain, (int, float)) and isinstance(aug_cost, (int, float)) and aug_cost > 0:
                return gain / aug_cost
        elif metric_id == "nf_count_per_aug_cost":
            nf_count = res.get("total_augmented_nf_count")
            aug_cost = res.get("security_augmentation_cost")
            if isinstance(nf_count, (int, float)) and isinstance(aug_cost, (int, float)) and aug_cost > 0:
                return nf_count / aug_cost
        elif metric_id == "avg_augmentation_gpu_per_node":
            gpu_units = res.get("total_augmentation_gpu_units")
            total_nodes = res.get("total_nodes")
            if isinstance(gpu_units, (int, float)) and isinstance(total_nodes, (int, float)) and total_nodes > 0:
                return gpu_units / total_nodes
        elif metric_id == "nodes_per_aug_cost":
            total_nodes = res.get("total_nodes")
            aug_cost = res.get("security_augmentation_cost")
            if isinstance(total_nodes, (int, float)) and isinstance(aug_cost, (int, float)) and aug_cost > 0:
                return total_nodes / aug_cost
        elif metric_id == "avg_clients_per_node":
            total_clients = res.get("total_clients")
            total_nodes = res.get("total_nodes")
            if isinstance(total_clients, (int, float)) and isinstance(total_nodes, (int, float)) and total_nodes > 0:
                return total_clients / total_nodes
        elif metric_id == "execution_time":
            return res.get("execution_time")
        elif metric_id == "total_nodes":
            return res.get("total_nodes")
        elif metric_id == "avg_cost_per_node":
            total_cost = res.get("total_cost")
            total_nodes = res.get("total_nodes")
            if isinstance(total_cost, (int, float)) and isinstance(total_nodes, (int, float)) and total_nodes > 0:
                return total_cost / total_nodes
    except Exception:  # noqa: BLE001
        return None
    return None


# ===== 1) 按节点数量聚合（compare_algorithms_by_nodes） ============================

def aggregate_by_nodes(all_results):
    """
    输入：compare_algorithms_by_nodes.run_all_algorithms_by_nodes() 的 all_results
    结构：all_results[node_size][security_requirement][degree][client_count][alg_name][node_nf_count] = [res, ...]

    返回：{node_size: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for node_size, results_by_security in all_results.items():
        per_size = {}
        for alg_name in ALGORITHM_NAMES:
            per_size[alg_name] = {}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                # 遍历所有 security_requirement / degree / client_count / node_nf_count / runs
                for security_results in results_by_security.values():
                    for degree_results in security_results.values():
                        for client_count, results_by_alg in degree_results.items():
                            nf_results = results_by_alg.get(alg_name, {})
                            for nf_count, run_list in nf_results.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_size[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_size[alg_name][metric_id] = (mn, mx, avg)

        summary[node_size] = per_size

    return summary


def write_nodes_sheet(wb, summary):
    ws = wb.create_sheet("按节点数量")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["节点数量", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for node_size in sorted(summary.keys()):
        per_size = summary[node_size]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"{node_size} 节点拓扑")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        row_idx = _write_metrics_block(ws, row_idx, node_size, per_size, header_fill, section_fill)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 2) 按 degree 聚合（compare_algorithms_by_degree） ===========================

def aggregate_by_degree(all_results):
    """
    all_results 来自 compare_algorithms_by_degree.run_all_algorithms_by_degree()
    结构：all_results[degree][security_requirement][node_size][client_count][alg_name][node_nf_count] = [res, ...]

    返回：{degree: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for degree, results_by_security in all_results.items():
        per_deg = {alg: {} for alg in ALGORITHM_NAMES}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                for security_results in results_by_security.values():
                    for size_results in security_results.values():
                        for client_count, results_by_alg in size_results.items():
                            nf_results = results_by_alg.get(alg_name, {})
                            for nf_count, run_list in nf_results.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_deg[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_deg[alg_name][metric_id] = (mn, mx, avg)

        summary[degree] = per_deg

    return summary


def write_degree_sheet(wb, summary):
    ws = wb.create_sheet("按度数")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["度数", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for degree in sorted(summary.keys()):
        per_deg = summary[degree]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"度数 {degree}")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        row_idx = _write_metrics_block(ws, row_idx, degree, per_deg, header_fill, section_fill)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 3) 按 client_count 聚合（compare_algorithms_by_clients） ====================

def aggregate_by_clients(all_results):
    """
    all_results 来自 compare_algorithms_by_clients.run_all_algorithms_by_clients()
    结构：all_results[client_count][security_requirement][node_size][degree][alg_name][node_nf_count] = [res, ...]

    返回：{client_count: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for client_count, results_by_security in all_results.items():
        per_client = {alg: {} for alg in ALGORITHM_NAMES}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                for security_results in results_by_security.values():
                    for size_results in security_results.values():
                        for degree_results in size_results.values():
                            nf_results = degree_results.get(alg_name, {})
                            for nf_count, run_list in nf_results.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_client[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_client[alg_name][metric_id] = (mn, mx, avg)

        summary[client_count] = per_client

    return summary


def write_clients_sheet(wb, summary):
    ws = wb.create_sheet("按客户端数")
    header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    section_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["客户端数", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for client_count in sorted(summary.keys()):
        per_client = summary[client_count]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"客户端数量 {client_count}")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        row_idx = _write_metrics_block(ws, row_idx, client_count, per_client, header_fill, section_fill)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 4) 按 NF instance 数聚合（compare_algorithms_by_nf_instances） ==============

def aggregate_by_nf(all_results):
    """
    all_results 来自 compare_algorithms_by_nf_instances.run_all_algorithms_by_nf_instances()
    结构：all_results[node_nf_count][security_requirement][node_size][degree][client_count][alg_name] = [res, ...]

    返回：{node_nf_count: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for node_nf_count, results_by_security in all_results.items():
        per_nf = {alg: {} for alg in ALGORITHM_NAMES}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                for security_results in results_by_security.values():
                    for size_results in security_results.values():
                        for degree_results in size_results.values():
                            for client_count, results_by_alg in degree_results.items():
                                run_list = results_by_alg.get(alg_name)
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_nf[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_nf[alg_name][metric_id] = (mn, mx, avg)

        summary[node_nf_count] = per_nf

    return summary


def write_nf_sheet(wb, summary):
    ws = wb.create_sheet("按NF个数")
    header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
    section_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["NF个数", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for nf_count in sorted(summary.keys()):
        per_nf = summary[nf_count]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"节点 NF 个数 {nf_count} (NODE_DATA_{nf_count})")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        row_idx = _write_metrics_block(ws, row_idx, nf_count, per_nf, header_fill, section_fill)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 5) 按安全需求聚合（compare_algorithms_by_security） =======================

def aggregate_by_security(all_results):
    """
    all_results 来自 compare_algorithms_by_security.run_all_algorithms_by_security_requirement()
    结构：all_results[security_requirement][node_size][degree][client_count][alg_name][node_nf_count] = [res, ...]

    返回：{security_requirement: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for security_requirement, results_by_size in all_results.items():
        per_security = {alg: {} for alg in ALGORITHM_NAMES}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                for size_results in results_by_size.values():
                    for degree_results in size_results.values():
                        for client_count, results_by_alg in degree_results.items():
                            nf_results = results_by_alg.get(alg_name, {})
                            for nf_count, run_list in nf_results.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_security[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_security[alg_name][metric_id] = (mn, mx, avg)

        summary[security_requirement] = per_security

    return summary


def write_security_sheet(wb, summary):
    ws = wb.create_sheet("按安全需求")
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    section_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["安全需求", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for security_requirement in sorted(summary.keys()):
        per_security = summary[security_requirement]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"安全需求 {security_requirement}")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        row_idx = _write_metrics_block(ws, row_idx, security_requirement, per_security, header_fill, section_fill)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 6) 按客户端接受率聚合（compare_algorithms_by_client_acceptance） ===========

def aggregate_by_client_acceptance(all_results):
    """
    all_results 来自 compare_algorithms_by_client_acceptance.run_all_algorithms_by_client_acceptance()
    结构：all_results[client_count][security_requirement][node_size][degree][alg_name][node_nf_count] = [res, ...]

    返回：{client_count: {alg_name: {"acceptance_rate": (min, max, avg)}}}
    注意：这个实验只输出接受率指标，不是标准的3个指标。
    """
    summary = {}

    for client_count, results_by_security in all_results.items():
        per_client = {alg: {} for alg in ALGORITHM_NAMES}

        for alg_name in ALGORITHM_NAMES:
            values = []
            # 遍历所有 security_requirement / node_size / degree / node_nf_count 组合
            for security_results in results_by_security.values():
                for size_results in security_results.values():
                    for degree_results in size_results.values():
                        nf_results_by_alg = degree_results.get(alg_name, {})
                        for nf_count, run_list in nf_results_by_alg.items():
                            if not isinstance(run_list, list):
                                continue
                            for res in run_list:
                                if not isinstance(res, dict):
                                    continue
                                served = res.get("total_clients")
                                if not isinstance(served, (int, float)):
                                    continue
                                # 接受率 = 实际服务客户端数 / 期望客户端数
                                acceptance = served / client_count if client_count else 0
                                values.append(acceptance)

            if not values:
                per_client[alg_name]["acceptance_rate"] = (None, None, None)
            else:
                mn = min(values)
                mx = max(values)
                avg = sum(values) / len(values)
                per_client[alg_name]["acceptance_rate"] = (mn, mx, avg)

        summary[client_count] = per_client

    return summary


def write_client_acceptance_sheet(wb, summary):
    ws = wb.create_sheet("按客户端接受率")
    header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    section_fill = PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    n_cols = 3 + len(ALGORITHM_NAMES)

    ws.append(["客户端数", "指标", "统计"] + ALGORITHM_NAMES)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "D2"

    row_idx = 2
    for client_count in sorted(summary.keys()):
        per_client = summary[client_count]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=1, value=f"客户端数量 {client_count}")
        cell.font = Font(bold=True, size=12)
        row_idx += 1
        # 分组标题
        ws.cell(row=row_idx, column=1, value=client_count)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
        cell = ws.cell(row=row_idx, column=2, value="【接受率】")
        cell.fill = section_fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1
        # 接受率 最小/最大/平均，指标名合并三行
        start_r = row_idx
        for stat_name, stat_index in (("最小值", 0), ("最大值", 1), ("平均值", 2)):
            ws.cell(row=row_idx, column=1, value=client_count)
            ws.cell(row=row_idx, column=2, value="接受率")
            ws.cell(row=row_idx, column=3, value=stat_name)
            for ci, alg_name in enumerate(ALGORITHM_NAMES):
                mn, mx, avg = per_client[alg_name].get("acceptance_rate", (None, None, None))
                val = (mn, mx, avg)[stat_index]
                cell = ws.cell(row=row_idx, column=4 + ci, value="{:.4f}".format(val) if isinstance(val, (int, float)) else "N/A")
                cell.alignment = Alignment(horizontal="right")
            row_idx += 1
        ws.merge_cells(start_row=start_r, start_column=2, end_row=row_idx - 1, end_column=2)
        ws.cell(row=start_r, column=2).alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


# ===== 主流程 =====================================================================

def main(
    output_path="results/六维实验对比汇总.xlsx",
    num_runs: int = DEFAULT_NUM_RUNS,
    topo_per_combo: int = DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO,
    verbose_progress: bool = DEFAULT_VERBOSE_PROGRESS,
):
    print("=" * 80)
    print("运行六个实验并导出 Excel 汇总")
    print("=" * 80)

    # 统一设置六个实验脚本的运行次数
    _set_all_num_runs(num_runs)
    # 统一设置每个 size/degree 生成的随机拓扑数量
    _set_all_generated_topologies_per_combo(topo_per_combo)
    # 统一设置是否打印详细进度信息
    _set_all_verbose_progress(verbose_progress)
    print(f"\n本次实验参数：NUM_RUNS = {num_runs}；每个(size, degree)生成的拓扑数量 = {topo_per_combo}；VERBOSE_PROGRESS = {verbose_progress}")

    # 1) 运行六个实验（只拿 all_results，不打印）
    print("\n[1/6] 运行按节点数量实验...")
    nodes_results = exp_nodes.run_all_algorithms_by_nodes()
    print("  ✓ 完成")

    print("\n[2/6] 运行按度数实验...")
    degree_results = exp_degree.run_all_algorithms_by_degree()
    print("  ✓ 完成")

    print("\n[3/6] 运行按客户端数量实验...")
    clients_results = exp_clients.run_all_algorithms_by_clients()
    print("  ✓ 完成")

    print("\n[4/6] 运行按NF个数实验...")
    nf_results = exp_nf.run_all_algorithms_by_nf_instances()
    print("  ✓ 完成")

    print("\n[5/6] 运行按安全需求实验...")
    security_results = exp_security.run_all_algorithms_by_security_requirement()
    print("  ✓ 完成")

    print("\n[6/6] 运行按客户端接受率实验...")
    acceptance_results = exp_client_acceptance.run_all_algorithms_by_client_acceptance()
    print("  ✓ 完成")

    # 2) 聚合
    print("\n开始聚合结果并写入 Excel ...")
    wb = Workbook()
    # 删除默认 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 按节点数量
    nodes_summary = aggregate_by_nodes(nodes_results)
    write_nodes_sheet(wb, nodes_summary)

    # 按度数
    degree_summary = aggregate_by_degree(degree_results)
    write_degree_sheet(wb, degree_summary)

    # 按客户端数量
    clients_summary = aggregate_by_clients(clients_results)
    write_clients_sheet(wb, clients_summary)

    # 按 NF instance 数
    nf_summary = aggregate_by_nf(nf_results)
    write_nf_sheet(wb, nf_summary)

    # 按安全需求
    security_summary = aggregate_by_security(security_results)
    write_security_sheet(wb, security_summary)

    # 按客户端接受率
    acceptance_summary = aggregate_by_client_acceptance(acceptance_results)
    write_client_acceptance_sheet(wb, acceptance_summary)

    # 3) 保存
    output_dir = os.path.dirname(output_path) or "results"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    wb.save(output_path)

    print(f"\n✓ 所有结果已导出到 Excel 文件: {output_path}")
    print("=" * 80)


if __name__ == "__main__":
    # 简单解析命令行参数：
    #   第一个参数（可选）：输出文件路径（默认 results/六维实验对比汇总.xlsx）
    #   第二个参数（可选）：NUM_RUNS（默认 DEFAULT_NUM_RUNS）
    #   第三个参数（可选）：每个 (size, degree) 生成的拓扑数量（默认 DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO）
    #   第四个参数（可选）：VERBOSE_PROGRESS（默认 DEFAULT_VERBOSE_PROGRESS，可用 1/true/True 或 0/false/False）
    default_output = os.path.join("results", "六维实验对比汇总.xlsx")

    out_path = default_output
    runs = DEFAULT_NUM_RUNS
    topo_per_combo = DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO
    verbose_progress = DEFAULT_VERBOSE_PROGRESS

    if len(sys.argv) >= 2 and sys.argv[1].strip():
        out_path = sys.argv[1]
    if len(sys.argv) >= 3:
        try:
            runs = int(sys.argv[2])
        except ValueError:
            print(f"警告: 无法解析运行次数 '{sys.argv[2]}', 使用默认值 {DEFAULT_NUM_RUNS}")
            runs = DEFAULT_NUM_RUNS
    if len(sys.argv) >= 4:
        try:
            topo_per_combo = int(sys.argv[3])
        except ValueError:
            print(
                f"警告: 无法解析拓扑数量 '{sys.argv[3]}', 使用默认值 {DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO}"
            )
            topo_per_combo = DEFAULT_GENERATED_TOPOLOGIES_PER_COMBO
    if len(sys.argv) >= 5:
        verbose_arg = sys.argv[4].strip().lower()
        if verbose_arg in ("1", "true", "yes", "y"):
            verbose_progress = True
        elif verbose_arg in ("0", "false", "no", "n"):
            verbose_progress = False
        else:
            print(
                f"警告: 无法解析 VERBOSE_PROGRESS '{sys.argv[4]}', 使用默认值 {DEFAULT_VERBOSE_PROGRESS}"
            )
            verbose_progress = DEFAULT_VERBOSE_PROGRESS

    main(out_path, num_runs=runs, topo_per_combo=topo_per_combo, verbose_progress=verbose_progress)


