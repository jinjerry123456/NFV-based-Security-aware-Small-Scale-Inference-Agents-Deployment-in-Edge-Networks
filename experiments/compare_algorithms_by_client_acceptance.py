# -*- coding: utf-8 -*-
"""
按客户端数量（10/20/30/40/50/60/70/80/90/100）评估四种算法的客户端接受率：
1. My Algorithm (基于Sec-MSL的贪心算法)
2. Cost-first Benchmark
3. Price-first Benchmark
4. Security-first Benchmark

自变量：client_count ∈ [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
聚合维度：
- 节点个数: 20, 30, 40, 50
- 拓扑度数: 2, 3, 4, 5
- 节点 NF instance 个数: NODE_DATA_2 ~ NODE_DATA_5
- 安全需求: 10, 18, 26, 34

核心指标：
- 客户端接受率 = 实际被服务客户端数 / 期望客户端数
  对每个 client_count 聚合最小值 / 最大值 / 平均值
"""

import os
import sys
import copy
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import importlib.util
from openpyxl import Workbook


# 导入四个算法
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 是否在批量实验中打印进度信息（大量 print 会拖慢 1000 次运行）
VERBOSE_PROGRESS = False


def _load_module(mod_name: str, rel_path: str):
    spec = importlib.util.spec_from_file_location(
        mod_name,
        os.path.join(base_dir, rel_path),
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


my_algorithm = _load_module("my_algorithm", "algorithms/my_algorithm.py")
benchmark_cost_first = _load_module("benchmark_cost_first", "algorithms/benchmark_cost-first.py")
benchmark_security_first = _load_module("benchmark_security_first", "algorithms/benchmark_security-first.py")
benchmark_price_first = _load_module("benchmark_price_first", "algorithms/benchmark_price-first.py")
# 为避免循环依赖，这里自定义聚合和写入函数（逻辑与 export_all_comparisons_to_excel 中一致）

ALGORITHM_NAMES = ["My Algorithm", "Cost-first", "Price-first", "Security-first"]

greedy_algorithm = my_algorithm.greedy_algorithm
cost_first_algorithm = benchmark_cost_first.cost_first_algorithm
price_first_algorithm = benchmark_price_first.price_first_algorithm
security_first_algorithm = benchmark_security_first.security_first_algorithm


# 导入拓扑与重置工具（改为动态生成拓扑）
import topology_config as topo
from topology_config import generate_multiple_topologies, reset_topology


# 不同 NF 个数对应的 NODE_DATA 模板名称
NODE_NF_COUNTS = [2, 3, 4, 5]

# 自变量：客户端数量（扩展到 10~100，每 10 个为一步）
CLIENT_COUNTS = [10, 20, 30, 40, 50]

# 聚合维度：安全需求值
SECURITY_REQUIREMENTS = [10, 18, 26, 34]

# 节点规模与度数
NODE_SIZES = [20, 30, 40, 50]
DEGREES = [2, 3, 4, 5]

# 每个配置下算法运行次数
NUM_RUNS = 20

# 每个 size/degree 生成的随机拓扑数量
GENERATED_TOPOLOGIES_PER_COMBO = 3
# 生成拓扑的基础随机种子（可调整以复现）
GENERATED_SEED_BASE = 2025


def run_algorithm_on_matrix_silently(algorithm_func, algorithm_name, adjacency_matrix, node_nf_count, client_count, security_requirement):
    """
    在给定邻接矩阵（2D）下静默运行算法。
    运行前挂载 client 数、安全需求、NODE_DATA_k；运行后恢复全局状态。
    """
    import sys as _sys
    from io import StringIO

    old_stdout = _sys.stdout
    buffer = StringIO()
    _sys.stdout = buffer

    # 备份需要修改的全局状态
    old_active_size = getattr(topo, "ACTIVE_TOPOLOGY_SIZE", None)
    old_get_active_adj = topo.get_active_adjacency_matrix
    old_client_count_by_topology = getattr(topo, "CLIENT_COUNT_BY_TOPOLOGY", None)
    old_security_requirement = getattr(topo, "SECURITY_REQUIREMENT", None)
    old_node_data = getattr(topo, "NODE_DATA", None)
    old_original_node_data = getattr(topo, "_ORIGINAL_NODE_DATA", None)

    try:
        size = len(adjacency_matrix)

        # 1) 配置该规模下的 client 数量，并清理相关缓存
        if not isinstance(topo.CLIENT_COUNT_BY_TOPOLOGY, dict):
            topo.CLIENT_COUNT_BY_TOPOLOGY = {}
        topo.CLIENT_COUNT_BY_TOPOLOGY[size] = client_count

        if hasattr(topo, "_CACHED_CLIENT_NODES"):
            try:
                topo._CACHED_CLIENT_NODES.pop(size, None)
            except Exception:  # noqa: BLE001
                pass
        if hasattr(topo, "_CACHED_CLIENT_IDS"):
            try:
                topo._CACHED_CLIENT_IDS.pop(size, None)
            except Exception:  # noqa: BLE001
                pass

        # 2) 设置全局安全需求值
        topo.SECURITY_REQUIREMENT = security_requirement

        # 3) 根据 node_nf_count 选择 NODE_DATA 模板
        node_var = f"NODE_DATA_{node_nf_count}"
        node_template = getattr(topo, node_var, None)
        if node_template is None:
            raise RuntimeError(f"topology_config 中未找到节点模板: {node_var}")

        topo.NODE_DATA = copy.deepcopy(node_template)
        try:
            topo._ORIGINAL_NODE_DATA = copy.deepcopy(topo.NODE_DATA)
        except Exception:  # noqa: BLE001
            pass

        # 4) 重置拓扑（恢复 NF、容量等到上述模板）
        reset_topology()

        # 5) 设置当前拓扑规模
        topo.ACTIVE_TOPOLOGY_SIZE = size

        # 6) 临时覆写 get_active_adjacency_matrix
        def _custom_get_active_adjacency_matrix():
            return adjacency_matrix

        topo.get_active_adjacency_matrix = _custom_get_active_adjacency_matrix

        # 7) 运行算法
        result = algorithm_func()

        _sys.stdout = old_stdout
        return result
    except Exception as e:  # noqa: BLE001
        _sys.stdout = old_stdout
        # print(f"✗ {algorithm_name} 在自定义拓扑上执行失败: {e}")
        # import traceback
        # traceback.print_exc()
        return None
    finally:
        # 8) 恢复 topo 全局状态
        topo.get_active_adjacency_matrix = old_get_active_adj
        if old_active_size is not None:
            topo.ACTIVE_TOPOLOGY_SIZE = old_active_size
        if old_client_count_by_topology is not None:
            topo.CLIENT_COUNT_BY_TOPOLOGY = old_client_count_by_topology
        if old_security_requirement is not None:
            topo.SECURITY_REQUIREMENT = old_security_requirement
        # 恢复 NODE_DATA 和 _ORIGINAL_NODE_DATA
        if old_node_data is not None:
            topo.NODE_DATA = old_node_data
        elif hasattr(topo, "NODE_DATA"):
            delattr(topo, "NODE_DATA")
        if old_original_node_data is not None:
            topo._ORIGINAL_NODE_DATA = old_original_node_data
        elif hasattr(topo, "_ORIGINAL_NODE_DATA"):
            topo._ORIGINAL_NODE_DATA = None


def run_all_algorithms_by_client_acceptance():
    """
    以 client_count ∈ [10..100] 为自变量，遍历 security_requirement / node_size / degree / NODE_DATA_2~5，
    运行四个算法，收集结果。

    Returns:
        dict: {
            client_count: {
                security_requirement: {
                    node_size: {
                        degree: {
                            algorithm_name: {
                                node_nf_count: [result_dict_or_None, ...]
                            }
                        }
                    }
                }
            }
        }
    """
    if VERBOSE_PROGRESS:
        print("=" * 80)
        print("按客户端数量对比四个算法的接受率 (client_count = 10..100)")
        print("聚合维度: security_requirement = 10/18/26/34, 节点个数, 度数, NODE_DATA_2~5")
        print("=" * 80)

    algorithm_configs = [
        ("My Algorithm", greedy_algorithm),
        ("Cost-first", cost_first_algorithm),
        ("Price-first", price_first_algorithm),
        ("Security-first", security_first_algorithm),
    ]

    all_results = {}

    for client_count in CLIENT_COUNTS:
        if VERBOSE_PROGRESS:
            print(f"\n{'=' * 80}")
            print(f"客户端数量: {client_count}")
            print(f"{'=' * 80}")

        per_client_results = {}

        for security_requirement in SECURITY_REQUIREMENTS:
            if VERBOSE_PROGRESS:
                print(f"\n  {'-' * 78}")
                print(f"  安全需求: {security_requirement}")
                print(f"  {'-' * 78}")

            per_security_results = {}

            for node_size in NODE_SIZES:
                if VERBOSE_PROGRESS:
                    print(f"\n  节点规模: {node_size}")

                per_size_results = {}

                for degree in DEGREES:
                    topologies = generate_multiple_topologies(
                        node_size,
                        degree,
                        count=GENERATED_TOPOLOGIES_PER_COMBO,
                        seed_base=GENERATED_SEED_BASE + node_size * 10 + degree,
                    )
                    if VERBOSE_PROGRESS:
                        print(f"    生成 {len(topologies)} 个随机拓扑 (degree={degree})")

                    per_degree_results = {name: {} for name, _ in algorithm_configs}

                    for node_nf_count in NODE_NF_COUNTS:
                        if VERBOSE_PROGRESS:
                            print(f"      使用节点 NF 个数配置: NODE_DATA_{node_nf_count}")
                        for idx, (alg_name, alg_func) in enumerate(algorithm_configs, 1):
                            all_runs = []
                            for topo_instance_id, matrix in topologies.items():
                                if VERBOSE_PROGRESS:
                                    print(
                                        f"        拓扑实例 {topo_instance_id}/{len(topologies)} -> 算法 {idx}/4: {alg_name}",
                                        end=" ... ",
                                    )
                                for run_idx in range(NUM_RUNS):
                                    res = run_algorithm_on_matrix_silently(
                                        alg_func,
                                        alg_name,
                                        matrix,
                                        node_nf_count=node_nf_count,
                                        client_count=client_count,
                                        security_requirement=security_requirement,
                                    )
                                    if isinstance(res, dict):
                                        res = dict(res)
                                        res["topology_size"] = node_size
                                        res["topology_degree"] = degree
                                        res["topology_instance_id"] = topo_instance_id
                                        res["security_requirement"] = security_requirement
                                    all_runs.append(res)
                                if VERBOSE_PROGRESS:
                                    print("✓")
                            per_degree_results[alg_name][node_nf_count] = all_runs

                    per_size_results[degree] = per_degree_results

                per_security_results[node_size] = per_size_results

            per_client_results[security_requirement] = per_security_results

        all_results[client_count] = per_client_results

    return all_results


def aggregate_by_client_acceptance(all_results):
    """
    结构：all_results[client_count][security_requirement][node_size][degree][alg_name][node_nf_count] = [res, ...]
    返回：{client_count: {alg_name: {"acceptance_rate": (min, max, avg)}}}
    """
    summary = {}

    for client_count, results_by_security in all_results.items():
        per_client = {alg: {} for alg in ALGORITHM_NAMES}

        for alg_name in ALGORITHM_NAMES:
            values = []
            for security_results in results_by_security.values():
                for size_results in security_results.values():
                    for degree_results in size_results.values():
                        nf_results_by_alg = degree_results.get(alg_name, {})
                        for nf_count, run_list in nf_results_by_alg.items():
                            if not isinstance(run_list, list):
                                continue
                            for res in run_list:
                                if not isinstance(res, dict):
                                    continue
                                served = res.get("total_clients")
                                if not isinstance(served, (int, float)):
                                    continue
                                acceptance = served / client_count if client_count else 0
                                values.append(acceptance)

            if not values:
                per_client[alg_name]["acceptance_rate"] = (None, None, None)
            else:
                mn = min(values)
                mx = max(values)
                avg = sum(values) / len(values)
                per_client[alg_name]["acceptance_rate"] = (mn, mx, avg)

        summary[client_count] = per_client

    return summary


def write_client_acceptance_sheet(wb, summary):
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet("按客户端接受率")

    header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    ws.append(["客户端数", "指标/统计"] + ALGORITHM_NAMES)
    for col in range(1, 2 + len(ALGORITHM_NAMES) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for client_count in sorted(summary.keys()):
        per_client = summary[client_count]

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(ALGORITHM_NAMES))
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"客户端数量 {client_count}"
        cell.font = Font(bold=True)
        row_idx += 1

        metric_name = "接受率"
        for stat_name, stat_index in (("最小值", 0), ("最大值", 1), ("平均值", 2)):
            row = [client_count, f"{metric_name}-{stat_name}"]
            for alg_name in ALGORITHM_NAMES:
                mn, mx, avg = per_client[alg_name].get("acceptance_rate", (None, None, None))
                val = (mn, mx, avg)[stat_index]
                if val is None:
                    row.append("N/A")
                else:
                    row.append("{:.4f}".format(val))
            ws.append(row)
            row_idx += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24


def print_summary(all_results):
    """
    按 client_count 聚合并输出客户端接受率的最小/最大/平均值。
    接受率 = 实际服务客户端数 / 期望客户端数。
    """
    print("\n" + "=" * 80)
    print("按客户端数量的算法接受率摘要")
    print("=" * 80)

    algorithm_names = ["My Algorithm", "Cost-first", "Price-first", "Security-first"]
    client_counts = sorted(all_results.keys())

    for client_count in client_counts:
        results_by_security = all_results[client_count]

        print(f"\n{'-' * 80}")
        print(
            f"客户端数量 {client_count}，在安全需求 {SECURITY_REQUIREMENTS}、"
            f"节点个数 {NODE_SIZES}、度数 {DEGREES} 与 NODE_DATA_2~5 上聚合"
        )
        print(f"{'-' * 80}")

        header = f"{'指标/统计':<20}"
        for alg_name in algorithm_names:
            header += f" {alg_name:<18}"
        print(header)
        print("-" * (20 + 19 * len(algorithm_names)))

        metric_name = "接受率"
        zero_cases = []  # 记录接受率为0的配置，便于调试
        for stat_name in ("最小值", "最大值", "平均值"):
            row = f"{metric_name}-{stat_name:<12}"
            for alg_name in algorithm_names:
                values = []
                # 遍历所有 security_requirement / node_size / degree / node_nf_count 组合
                for security_results in results_by_security.values():
                    for size_results in security_results.values():
                        for degree_results in size_results.values():
                            nf_results_by_alg = degree_results.get(alg_name, {})
                            for nf_count, run_list in nf_results_by_alg.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    if not isinstance(res, dict):
                                        continue
                                    served = res.get("total_clients")
                                    if not isinstance(served, (int, float)):
                                        continue
                                    # 接受率 = 实际服务客户端数 / 期望客户端数
                                    acceptance = served / client_count if client_count else 0
                                    values.append(acceptance)
                                    if acceptance == 0:
                                        zero_cases.append(
                                            {
                                                "client_count": client_count,
                                                "security_requirement": res.get("security_requirement"),
                                                "node_size": res.get("topology_size"),
                                                "degree": res.get("topology_degree"),
                                                "node_nf_count": nf_count,
                                                "algorithm": alg_name,
                                            }
                                        )

                if not values:
                    row += f" {'N/A':<18}"
                    continue

                if stat_name == "最小值":
                    agg_val = min(values)
                elif stat_name == "最大值":
                    agg_val = max(values)
                else:
                    agg_val = sum(values) / len(values)

                row += f" {agg_val:<18.4f}"
            print(row)

        # 如果出现接受率为0的配置，打印调试信息
        if zero_cases:
            print("  [DEBUG] 以下组合接受率为0（只列出前20条，更多请加日志或导出）：")
            for case in zero_cases[:20]:
                print(
                    f"    alg={case['algorithm']}, clients={case['client_count']}, "
                    f"sec={case['security_requirement']}, size={case['node_size']}, "
                    f"degree={case['degree']}, NODE_DATA_{case['node_nf_count']}"
                )
            if len(zero_cases) > 20:
                print(f"    ... 还有 {len(zero_cases) - 20} 条未显示 ...")

    print("=" * 80)


def main():
    print("=" * 80)
    print("按客户端数量（10..100）对比四个算法的接受率")
    print("=" * 80)

    all_results = run_all_algorithms_by_client_acceptance()
    print_summary(all_results)

    # 聚合并写入 Excel，命名规则：client_acceptance + GENERATED_TOPOLOGIES_PER_COMBO + GENERATED_SEED_BASE + time
    print("\n开始聚合并写入 Excel ...")
    summary = aggregate_by_client_acceptance(all_results)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    write_client_acceptance_sheet(wb, summary)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_name = (
        f"client_acceptance_{NUM_RUNS}_{GENERATED_TOPOLOGIES_PER_COMBO}_"
        f"{timestamp}.xlsx"
    )
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "results")
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, excel_name)
    wb.save(excel_path)

    print(f"✓ 已保存 Excel: {excel_path}")


if __name__ == "__main__":
    main()

