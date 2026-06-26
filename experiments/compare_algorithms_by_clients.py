# -*- coding: utf-8 -*-
"""
按客户端数量（client_count ∈ [3, 5, 8, 10]）对比四种算法：
1. My Algorithm (基于Sec-MSL的贪心算法)
2. Cost-first Benchmark
3. Price-first Benchmark
4. Security-first Benchmark

自变量：客户端数量 client_count

在每个 client_count 下聚合的维度：
- 节点个数: [20, 30, 40, 50]
- 拓扑度数: degree ∈ [2, 3, 4, 5]
- 节点 NF instance 个数: NODE_DATA_2, NODE_DATA_3, NODE_DATA_4, NODE_DATA_5

输出三个指标在上述所有组合上的：
- 最小值
- 最大值
- 平均值

指标：
1. 总体成本 / 平均安全级别 (overall cost per security level)
2. 平均安全级别 (average security)
3. 总体成本 (安全开销 + 算力开销 + 带宽开销)
"""

import os
import sys
import copy
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import importlib.util
from openpyxl import Workbook


# 导入四个算法
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 是否在批量实验中打印进度信息（大量 print 会拖慢 1000 次运行）
VERBOSE_PROGRESS = False


def _load_module(mod_name: str, rel_path: str):
    spec = importlib.util.spec_from_file_location(
        mod_name,
        os.path.join(base_dir, rel_path),
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


my_algorithm = _load_module("my_algorithm", "algorithms/my_algorithm.py")
benchmark_cost_first = _load_module("benchmark_cost_first", "algorithms/benchmark_cost-first.py")
benchmark_security_first = _load_module("benchmark_security_first", "algorithms/benchmark_security-first.py")
benchmark_price_first = _load_module("benchmark_price_first", "algorithms/benchmark_price-first.py")

greedy_algorithm = my_algorithm.greedy_algorithm
cost_first_algorithm = benchmark_cost_first.cost_first_algorithm
price_first_algorithm = benchmark_price_first.price_first_algorithm
security_first_algorithm = benchmark_security_first.security_first_algorithm

# 供聚合/写表使用的公共定义
ALGORITHM_NAMES = ["My Algorithm", "Cost-first", "Price-first", "Security-first"]
# (展示名称, 内部ID, 格式化字符串)
METRICS = [
    ("总体成本/安全级别", "overall_cost_per_security", "{:.4f}"),
    ("平均安全级别", "avg_security", "{:.2f}"),
    ("总体成本 ($)", "overall_cost", "{:.2f}"),
    ("运行时间 (s)", "execution_time", "{:.4f}"),
]


# 导入拓扑与重置工具（改为动态生成拓扑）
import topology_config as topo
from topology_config import generate_multiple_topologies, reset_topology


# 不同 NF 个数对应的 NODE_DATA 模板名称
NODE_NF_COUNTS = [2, 3, 4, 5]

# 自变量：不同实验中要测试的 client 数量
CLIENT_COUNTS = [3, 5, 8, 10]

# 聚合维度：不同实验中要测试的安全需求值
SECURITY_REQUIREMENTS = [10, 18, 26, 34]

# 要考虑的节点规模
NODE_SIZES = [20, 30, 40, 50]

# 每个配置下算法运行次数
NUM_RUNS = 100

# 每个 size/degree 生成的随机拓扑数量
GENERATED_TOPOLOGIES_PER_COMBO = 3
# 生成拓扑的基础随机种子（可调整以复现）
GENERATED_SEED_BASE = 2025


def run_algorithm_on_matrix_silently(algorithm_func, algorithm_name, adjacency_matrix, node_nf_count, client_count, security_requirement):
    """
    在给定邻接矩阵（2D）下静默运行算法。

    做的事情：
    1. 为给定拓扑规模 size，设置每次实际服务的 client 数量（CLIENT_COUNT_BY_TOPOLOGY[size] = client_count），并清缓存；
    2. 设置全局安全需求值（SECURITY_REQUIREMENT = security_requirement）；
    3. 将 topo.NODE_DATA 设置为指定 NF 个数的模板（NODE_DATA_2/3/4/5），并同步 _ORIGINAL_NODE_DATA；
    4. reset_topology() 恢复该模板为干净状态；
    5. 暂时设置 topo.ACTIVE_TOPOLOGY_SIZE = 节点数；
    6. 暂时覆写 topo.get_active_adjacency_matrix，使其返回给定矩阵；
    7. 运行算法函数并捕获 stdout；
    8. 最后恢复被改动的全局状态。
    """
    import sys as _sys
    from io import StringIO

    old_stdout = _sys.stdout
    buffer = StringIO()
    _sys.stdout = buffer

    # 备份需要修改的全局状态
    old_active_size = getattr(topo, "ACTIVE_TOPOLOGY_SIZE", None)
    old_get_active_adj = topo.get_active_adjacency_matrix
    old_client_count_by_topology = getattr(topo, "CLIENT_COUNT_BY_TOPOLOGY", None)
    old_security_requirement = getattr(topo, "SECURITY_REQUIREMENT", None)
    old_node_data = getattr(topo, "NODE_DATA", None)
    old_original_node_data = getattr(topo, "_ORIGINAL_NODE_DATA", None)

    try:
        # 0) 计算拓扑规模（节点数）
        size = len(adjacency_matrix)

        # 1) 配置该规模下的 client 数量，并清理相关缓存
        if not isinstance(topo.CLIENT_COUNT_BY_TOPOLOGY, dict):
            topo.CLIENT_COUNT_BY_TOPOLOGY = {}
        topo.CLIENT_COUNT_BY_TOPOLOGY[size] = client_count

        if hasattr(topo, "_CACHED_CLIENT_NODES"):
            try:
                topo._CACHED_CLIENT_NODES.pop(size, None)
            except Exception:  # noqa: BLE001
                pass
        if hasattr(topo, "_CACHED_CLIENT_IDS"):
            try:
                topo._CACHED_CLIENT_IDS.pop(size, None)
            except Exception:  # noqa: BLE001
                pass

        # 2) 设置全局安全需求值
        topo.SECURITY_REQUIREMENT = security_requirement

        # 3) 根据 node_nf_count 选择 NODE_DATA 模板
        node_var = f"NODE_DATA_{node_nf_count}"
        node_template = getattr(topo, node_var, None)
        if node_template is None:
            raise RuntimeError(f"topology_config 中未找到节点模板: {node_var}")

        topo.NODE_DATA = copy.deepcopy(node_template)
        try:
            topo._ORIGINAL_NODE_DATA = copy.deepcopy(topo.NODE_DATA)
        except Exception:  # noqa: BLE001
            pass

        # 4) 重置拓扑（恢复 NF、容量等到上述模板）
        reset_topology()

        # 5) 设置当前拓扑规模
        topo.ACTIVE_TOPOLOGY_SIZE = size

        # 6) 临时覆写 get_active_adjacency_matrix
        def _custom_get_active_adjacency_matrix():
            return adjacency_matrix

        topo.get_active_adjacency_matrix = _custom_get_active_adjacency_matrix

        # 7) 运行算法
        result = algorithm_func()

        # 恢复 stdout
        _sys.stdout = old_stdout
        return result
    except Exception as e:  # noqa: BLE001
        _sys.stdout = old_stdout
        print(f"✗ {algorithm_name} 在自定义拓扑上执行失败: {e}")
        import traceback

        traceback.print_exc()
        return None
    finally:
        # 8) 恢复 topo 全局状态
        topo.get_active_adjacency_matrix = old_get_active_adj
        if old_active_size is not None:
            topo.ACTIVE_TOPOLOGY_SIZE = old_active_size
        if old_client_count_by_topology is not None:
            topo.CLIENT_COUNT_BY_TOPOLOGY = old_client_count_by_topology
        if old_security_requirement is not None:
            topo.SECURITY_REQUIREMENT = old_security_requirement
        # 恢复 NODE_DATA 和 _ORIGINAL_NODE_DATA
        if old_node_data is not None:
            topo.NODE_DATA = old_node_data
        elif hasattr(topo, "NODE_DATA"):
            delattr(topo, "NODE_DATA")
        if old_original_node_data is not None:
            topo._ORIGINAL_NODE_DATA = old_original_node_data
        elif hasattr(topo, "_ORIGINAL_NODE_DATA"):
            topo._ORIGINAL_NODE_DATA = None


def run_all_algorithms_by_clients():
    """
    以 client_count ∈ [3, 5, 8, 10] 作为自变量：
    - 对每个 client_count：
      - 遍历安全需求 security_requirement ∈ {10, 20, 30, 40}；
      - 遍历节点规模 size ∈ {20,30,40,50}；
      - 遍历 degree ∈ {2,3,4,5} 对应的 NETWORK_TOPOLOGY_size_degreeD；
      - 遍历 NODE_DATA_2~5；
      - 运行四个算法并收集结果。

    Returns:
        dict: {
            client_count: {
                security_requirement: {
                    node_size: {
                        degree: {
                            algorithm_name: {node_nf_count: [result_dict_or_None, ... 共 NUM_RUNS 个]}
                        }
                    }
                }
            }
        }
    """
    if VERBOSE_PROGRESS:
        print("=" * 80)
        print("按客户端数量对比四个算法 (client_count = 3/5/8/10)")
        print("聚合维度: security_requirement = 10/20/30/40, 节点个数, 度数, NODE_DATA")
        print("=" * 80)

    algorithm_configs = [
        ("My Algorithm", greedy_algorithm),
        ("Cost-first", cost_first_algorithm),
        ("Price-first", price_first_algorithm),
        ("Security-first", security_first_algorithm),
    ]

    all_results = {}
    degrees = [2, 3, 4, 5]

    for client_count in CLIENT_COUNTS:
        if VERBOSE_PROGRESS:
            print(f"\n{'=' * 80}")
            print(f"客户端数量: {client_count}")
            print(f"{'=' * 80}")

        # per_client_results[security_requirement][node_size][degree][算法名][node_nf_count] = [result, ... NUM_RUNS]
        per_client_results = {}

        for security_requirement in SECURITY_REQUIREMENTS:
            if VERBOSE_PROGRESS:
                print(f"\n  {'-' * 78}")
                print(f"  安全需求: {security_requirement}")
                print(f"  {'-' * 78}")

            # per_security_results[node_size][degree][算法名][node_nf_count] = [result, ... NUM_RUNS]
            per_security_results = {}

            for node_size in NODE_SIZES:
                if VERBOSE_PROGRESS:
                    print(f"\n  节点规模: {node_size}")

                per_size_results = {}

                for degree in degrees:
                    # 为每个 size/degree 生成多份随机拓扑
                    topologies = generate_multiple_topologies(
                        node_size,
                        degree,
                        count=GENERATED_TOPOLOGIES_PER_COMBO,
                        seed_base=GENERATED_SEED_BASE + node_size * 10 + degree,
                    )
                    if VERBOSE_PROGRESS:
                        print(f"    使用拓扑度数: {degree}，生成 {len(topologies)} 个随机拓扑")

                    per_degree_results = {name: {} for name, _ in algorithm_configs}

                    for node_nf_count in NODE_NF_COUNTS:
                        if VERBOSE_PROGRESS:
                            print(f"      使用节点 NF 个数配置: NODE_DATA_{node_nf_count}")
                        for idx, (alg_name, alg_func) in enumerate(algorithm_configs, 1):
                            all_runs = []
                            for topo_instance_id, matrix in topologies.items():
                                if VERBOSE_PROGRESS:
                                    print(
                                        f"        拓扑实例 {topo_instance_id}/{len(topologies)} -> 算法 {idx}/4: {alg_name}",
                                        end=" ... ",
                                    )
                                for run_idx in range(NUM_RUNS):
                                    res = run_algorithm_on_matrix_silently(
                                        alg_func,
                                        alg_name,
                                        matrix,
                                        node_nf_count=node_nf_count,
                                        client_count=client_count,
                                        security_requirement=security_requirement,
                                    )
                                    # 标记拓扑信息，便于后续聚合/排查
                                    if isinstance(res, dict):
                                        res = dict(res)
                                        res["topology_size"] = node_size
                                        res["topology_degree"] = degree
                                        res["topology_instance_id"] = topo_instance_id
                                        res["security_requirement"] = security_requirement
                                    all_runs.append(res)
                                if VERBOSE_PROGRESS:
                                    print("✓")
                            per_degree_results[alg_name][node_nf_count] = all_runs

                    per_size_results[degree] = per_degree_results

                per_security_results[node_size] = per_size_results

            per_client_results[security_requirement] = per_security_results

        all_results[client_count] = per_client_results

    return all_results


def print_summary(all_results):
    """
    打印按 client_count 聚合后的算法对比摘要：
    - 自变量：client_count = 3/5/8/10
    - 在每个 client_count 下，对所有 security_requirement / node_size / degree / NODE_DATA_2~5 组合聚合；
    - 对每个指标输出：最小值 / 最大值 / 平均值。
    """
    print("\n" + "=" * 80)
    print("按客户端数量的算法对比摘要")
    print("=" * 80)

    algorithm_names = ALGORITHM_NAMES
    client_counts = sorted(all_results.keys())

    metrics = METRICS

    for client_count in client_counts:
        results_by_security = all_results[client_count]

        print(f"\n{'-' * 80}")
        print(
            f"客户端数量 {client_count}，在安全需求 {SECURITY_REQUIREMENTS}、"
            f"节点个数 {NODE_SIZES}、度数 2/3/4/5 与 NODE_DATA_2~5 上聚合"
        )
        print(f"{'-' * 80}")

        header = f"{'指标/统计':<20}"
        for alg_name in algorithm_names:
            header += f" {alg_name:<18}"
        print(header)
        print("-" * (20 + 19 * len(algorithm_names)))

        for metric_name, metric_id, fmt in metrics:
            for stat_name in ("最小值", "最大值", "平均值"):
                row = f"{metric_name}-{stat_name:<12}"
                for alg_name in algorithm_names:
                    values = []
                    # 遍历所有 security_requirement / node_size / degree / node_nf_count 组合
                    for security_results in results_by_security.values():
                        for size_results in security_results.values():
                            for degree_results in size_results.values():
                                nf_results_by_alg = degree_results.get(alg_name, {})
                                for nf_count, run_list in nf_results_by_alg.items():
                                    if not isinstance(run_list, list):
                                        continue
                                    for res in run_list:
                                        if not isinstance(res, dict):
                                            continue
                                        val = None
                                        try:
                                            if metric_id == "overall_cost_per_security":
                                                total_cost = res.get("total_cost")
                                                avg_sec = res.get("avg_node_security_level")
                                                if isinstance(total_cost, (int, float)) and isinstance(avg_sec, (int, float)) and avg_sec > 0:
                                                    val = total_cost / avg_sec
                                            elif metric_id == "avg_security":
                                                val = res.get("avg_node_security_level")
                                            elif metric_id == "overall_cost":
                                                val = res.get("total_cost")
                                            elif metric_id == "execution_time":
                                                val = res.get("execution_time")
                                        except Exception:  # noqa: BLE001
                                            val = None

                                        if isinstance(val, (int, float)):
                                            values.append(val)

                    if not values:
                        row += f" {'N/A':<18}"
                        continue

                    if stat_name == "最小值":
                        agg_val = min(values)
                    elif stat_name == "最大值":
                        agg_val = max(values)
                    else:
                        agg_val = sum(values) / len(values)

                    if fmt == "{}":
                        row += f" {fmt.format(int(agg_val)):<18}"
                    else:
                        row += f" {fmt.format(agg_val):<18}"
                print(row)

    print("=" * 80)


def _extract_metric_from_result(metric_id, res):
    """从单次算法结果字典中提取一个指标值。"""
    if not isinstance(res, dict):
        return None
    try:
        if metric_id == "overall_cost_per_security":
            total_cost = res.get("total_cost")
            avg_sec = res.get("avg_node_security_level")
            if isinstance(total_cost, (int, float)) and isinstance(avg_sec, (int, float)) and avg_sec > 0:
                return total_cost / avg_sec
        elif metric_id == "avg_security":
            return res.get("avg_node_security_level")
        elif metric_id == "overall_cost":
            return res.get("total_cost")
        elif metric_id == "execution_time":
            return res.get("execution_time")
    except Exception:  # noqa: BLE001
        return None
    return None


def aggregate_by_clients(all_results):
    """
    结构：all_results[client_count][security_requirement][node_size][degree][alg_name][node_nf_count] = [res, ...]
    返回：{client_count: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for client_count, results_by_security in all_results.items():
        per_client = {alg: {} for alg in ALGORITHM_NAMES}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                for security_results in results_by_security.values():
                    for size_results in security_results.values():
                        for degree_results in size_results.values():
                            nf_results_by_alg = degree_results.get(alg_name, {})
                            for nf_count, run_list in nf_results_by_alg.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_client[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_client[alg_name][metric_id] = (mn, mx, avg)

        summary[client_count] = per_client

    return summary


def write_clients_sheet(wb, summary):
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet("按客户端数")

    header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    ws.append(["客户端数", "指标/统计"] + ALGORITHM_NAMES)
    for col in range(1, 2 + len(ALGORITHM_NAMES) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for client_count in sorted(summary.keys()):
        per_client = summary[client_count]

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(ALGORITHM_NAMES))
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"客户端数量 {client_count}"
        cell.font = Font(bold=True)
        row_idx += 1

        for metric_name, metric_id, fmt in METRICS:
            for stat_name, stat_index in (("最小值", 0), ("最大值", 1), ("平均值", 2)):
                row = [client_count, f"{metric_name}-{stat_name}"]
                for alg_name in ALGORITHM_NAMES:
                    mn, mx, avg = per_client[alg_name].get(metric_id, (None, None, None))
                    val = (mn, mx, avg)[stat_index]
                    if val is None:
                        row.append("N/A")
                    else:
                        row.append(fmt.format(val))
                ws.append(row)
                row_idx += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24


def main():
    print("=" * 80)
    print("按客户端数量（3/5/8/10）对比四个算法")
    print("=" * 80)

    all_results = run_all_algorithms_by_clients()
    print_summary(all_results)

    # 聚合并写入 Excel，命名规则：client_count + GENERATED_TOPOLOGIES_PER_COMBO + GENERATED_SEED_BASE + time
    print("\n开始聚合并写入 Excel ...")
    summary = aggregate_by_clients(all_results)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    write_clients_sheet(wb, summary)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_name = (
        f"client_count_{NUM_RUNS}_{GENERATED_TOPOLOGIES_PER_COMBO}_"
        f"{timestamp}.xlsx"
    )
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "results")
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, excel_name)
    wb.save(excel_path)

    print(f"✓ 已保存 Excel: {excel_path}")


if __name__ == "__main__":
    main()


