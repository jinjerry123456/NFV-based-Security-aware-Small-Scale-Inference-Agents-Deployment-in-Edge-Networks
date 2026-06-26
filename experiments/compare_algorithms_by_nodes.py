# -*- coding: utf-8 -*-
"""
按节点数量（20/30/40/50）对比四种算法：
1. My Algorithm (基于Sec-MSL的贪心算法)
2. Cost-first Benchmark
3. Price-first Benchmark
4. Security-first Benchmark

这里采用拓扑矩阵：
- 20 节点: NETWORK_TOPOLOGY_20_2D
- 30 节点: NETWORK_TOPOLOGY_30_2D
- 40 节点: NETWORK_TOPOLOGY_40_2D
- 50 节点: NETWORK_TOPOLOGY_50_2D

核心思路：
- 不修改算法代码，只在运行前临时“挂载”对应规模的邻接矩阵，
  并设置 ACTIVE_TOPOLOGY_SIZE = 节点数；
- 运行完一个组合后，恢复 topology_config 的原始状态；
- 打印各算法在不同节点数量下的关键指标，方便做对比。
"""

import os
import sys
import copy
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import importlib.util
from openpyxl import Workbook


# 导入四个算法（与 compare_algorithms.py 保持一致的方式）
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 是否在批量实验中打印进度信息（大量 print 会拖慢 1000 次运行）
VERBOSE_PROGRESS = False


def _load_module(mod_name: str, rel_path: str):
    spec = importlib.util.spec_from_file_location(
        mod_name,
        os.path.join(base_dir, rel_path),
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


my_algorithm = _load_module("my_algorithm", "algorithms/my_algorithm.py")
benchmark_cost_first = _load_module("benchmark_cost_first", "algorithms/benchmark_cost-first.py")
benchmark_security_first = _load_module("benchmark_security_first", "algorithms/benchmark_security-first.py")
benchmark_price_first = _load_module("benchmark_price_first", "algorithms/benchmark_price-first.py")

greedy_algorithm = my_algorithm.greedy_algorithm
cost_first_algorithm = benchmark_cost_first.cost_first_algorithm
price_first_algorithm = benchmark_price_first.price_first_algorithm
security_first_algorithm = benchmark_security_first.security_first_algorithm


# 导入拓扑与重置工具（改为动态生成拓扑）
import topology_config as topo
from topology_config import generate_multiple_topologies, reset_topology

# 不同 NF 个数对应的 NODE_DATA 模板名称
NODE_NF_COUNTS = [2, 3, 4, 5]

# 不同实验中要测试的 client 数量（每个拓扑规模都测试这几种）
CLIENT_COUNTS = [3, 5, 8, 10]

# 聚合维度：不同实验中要测试的安全需求值
SECURITY_REQUIREMENTS = [10, 18, 26, 34]

# 每个配置下算法运行次数
NUM_RUNS = 100

# 要考虑的节点规模
NODE_SIZES = [20, 30, 40, 50]

# 要考虑的拓扑度数
DEGREES = [2, 3, 4, 5]

# 每个 size/degree 生成的随机拓扑数量
GENERATED_TOPOLOGIES_PER_COMBO = 3
# 生成拓扑的基础随机种子（可调整以复现）
GENERATED_SEED_BASE = 2025

# 聚合/写表时使用的公共定义
ALGORITHM_NAMES = ["My Algorithm", "Cost-first", "Price-first", "Security-first"]
# (展示名称, 内部ID, 格式化字符串)
METRICS = [
    ("总体成本/安全级别", "overall_cost_per_security", "{:.4f}"),
    ("平均安全级别", "avg_security", "{:.2f}"),
    ("总体成本 ($)", "overall_cost", "{:.2f}"),
    ("运行时间 (s)", "execution_time", "{:.4f}"),
]


def run_algorithm_on_matrix_silently(algorithm_func, algorithm_name, adjacency_matrix, node_nf_count, client_count, security_requirement):
    """
    在给定邻接矩阵（2D）下静默运行算法。

    做的事情：
    1. 为给定拓扑规模 size，设置每次实际服务的 client 数量（CLIENT_COUNT_BY_TOPOLOGY[size] = client_count），并清缓存；
    2. 设置全局安全需求值（SECURITY_REQUIREMENT = security_requirement）；
    3. 将 topo.NODE_DATA 设置为指定 NF 个数的模板（NODE_DATA_2/3/4/5），并同步 _ORIGINAL_NODE_DATA；
    4. reset_topology() 恢复该模板为干净状态；
    5. 暂时设置 topo.ACTIVE_TOPOLOGY_SIZE = 节点数；
    6. 暂时覆写 topo.get_active_adjacency_matrix，使其返回给定矩阵；
    7. 运行算法函数并捕获 stdout；
    8. 最后恢复被改动的全局状态。
    """
    import sys as _sys
    from io import StringIO

    old_stdout = _sys.stdout
    buffer = StringIO()
    _sys.stdout = buffer

    # 备份需要修改的全局状态
    old_active_size = getattr(topo, "ACTIVE_TOPOLOGY_SIZE", None)
    old_get_active_adj = topo.get_active_adjacency_matrix
    old_client_count_by_topology = getattr(topo, "CLIENT_COUNT_BY_TOPOLOGY", None)
    old_security_requirement = getattr(topo, "SECURITY_REQUIREMENT", None)
    old_node_data = getattr(topo, "NODE_DATA", None)
    old_original_node_data = getattr(topo, "_ORIGINAL_NODE_DATA", None)

    try:
        # 0) 计算拓扑规模（节点数）
        size = len(adjacency_matrix)

        # 1) 配置该规模下的 client 数量，并清理相关缓存
        if not isinstance(topo.CLIENT_COUNT_BY_TOPOLOGY, dict):
            topo.CLIENT_COUNT_BY_TOPOLOGY = {}
        topo.CLIENT_COUNT_BY_TOPOLOGY[size] = client_count

        if hasattr(topo, "_CACHED_CLIENT_NODES"):
            try:
                topo._CACHED_CLIENT_NODES.pop(size, None)
            except Exception:  # noqa: BLE001
                pass
        if hasattr(topo, "_CACHED_CLIENT_IDS"):
            try:
                topo._CACHED_CLIENT_IDS.pop(size, None)
            except Exception:  # noqa: BLE001
                pass

        # 2) 设置全局安全需求值
        topo.SECURITY_REQUIREMENT = security_requirement

        # 3) 根据 node_nf_count 选择 NODE_DATA 模板
        node_var = f"NODE_DATA_{node_nf_count}"
        node_template = getattr(topo, node_var, None)
        if node_template is None:
            raise RuntimeError(f"topology_config 中未找到节点模板: {node_var}")

        topo.NODE_DATA = copy.deepcopy(node_template)
        try:
            topo._ORIGINAL_NODE_DATA = copy.deepcopy(topo.NODE_DATA)
        except Exception:  # noqa: BLE001
            pass

        # 4) 重置拓扑（恢复 NF、容量等到上述模板）
        reset_topology()

        # 5) 设置当前拓扑规模
        topo.ACTIVE_TOPOLOGY_SIZE = size

        # 6) 临时覆写 get_active_adjacency_matrix
        def _custom_get_active_adjacency_matrix():
            return adjacency_matrix

        topo.get_active_adjacency_matrix = _custom_get_active_adjacency_matrix

        # 7) 运行算法
        result = algorithm_func()

        # 恢复 stdout
        _sys.stdout = old_stdout
        return result
    except Exception as e:  # noqa: BLE001
        _sys.stdout = old_stdout
        print(f"✗ {algorithm_name} 在自定义拓扑上执行失败: {e}")
        import traceback

        traceback.print_exc()
        return None
    finally:
        # 8) 恢复 topo 全局状态
        topo.get_active_adjacency_matrix = old_get_active_adj
        if old_active_size is not None:
            topo.ACTIVE_TOPOLOGY_SIZE = old_active_size
        if old_client_count_by_topology is not None:
            topo.CLIENT_COUNT_BY_TOPOLOGY = old_client_count_by_topology
        if old_security_requirement is not None:
            topo.SECURITY_REQUIREMENT = old_security_requirement
        # 恢复 NODE_DATA 和 _ORIGINAL_NODE_DATA
        if old_node_data is not None:
            topo.NODE_DATA = old_node_data
        elif hasattr(topo, "NODE_DATA"):
            delattr(topo, "NODE_DATA")
        if old_original_node_data is not None:
            topo._ORIGINAL_NODE_DATA = old_original_node_data
        elif hasattr(topo, "_ORIGINAL_NODE_DATA"):
            topo._ORIGINAL_NODE_DATA = None


def run_all_algorithms_by_nodes():
    """
    在 20/30/40/50 四种节点规模上，
    对每个安全需求 (10/18/26/34)，
    对每个度数 (2/3/4/5) 生成多种随机拓扑，
    在每种拓扑下，对 client 数量 (3/5/8/10) + NF 个数配置（NODE_DATA_2/3/4/5）
    运行四个算法，并返回结果。

    Returns:
        dict: {
            node_size: {
                security_requirement: {
                degree: {
                    client_count: {
                        algorithm_name: {
                            node_nf_count: [result_dict_or_None, ... 共 NUM_RUNS * 拓扑数量 个]
                            }
                        }
                    }
                }
            }
        }
    """
    if VERBOSE_PROGRESS:
        print("=" * 80)
        print("按节点数量对比四个算法 (使用随机生成的拓扑)")
        print("聚合维度: security_requirement = 10/18/26/34, 度数, 客户端数量, NODE_DATA")
        print("=" * 80)

    algorithm_configs = [
        ("My Algorithm", greedy_algorithm),
        ("Cost-first", cost_first_algorithm),
        ("Price-first", price_first_algorithm),
        ("Security-first", security_first_algorithm),
    ]

    all_results = {}

    for node_size in NODE_SIZES:
        if VERBOSE_PROGRESS:
            print(f"\n{'=' * 80}")
            print(f"节点规模: {node_size}")
            print(f"{'=' * 80}")

        # per_size_results[security_requirement][degree][client_count][算法名][node_nf_count] = [result, ...]
        per_size_results = {}

        for security_requirement in SECURITY_REQUIREMENTS:
            if VERBOSE_PROGRESS:
                print(f"\n  {'-' * 78}")
                print(f"  安全需求: {security_requirement}")
                print(f"  {'-' * 78}")

            # per_security_results[degree][client_count][算法名][node_nf_count] = [result, ...]
            per_security_results = {}

            for degree in DEGREES:
                # 为每个 size/degree 生成多份随机拓扑
                topologies = generate_multiple_topologies(
                    node_size,
                    degree,
                    count=GENERATED_TOPOLOGIES_PER_COMBO,
                    seed_base=GENERATED_SEED_BASE + node_size * 10 + degree,
                )
                if VERBOSE_PROGRESS:
                    print(f"\n  使用拓扑度数: {degree}，生成 {len(topologies)} 个随机拓扑")

                per_degree_results = {}

                for client_count in CLIENT_COUNTS:
                    if VERBOSE_PROGRESS:
                        print(f"    使用客户端数量: {client_count}")
                    per_client_results = {name: {} for name, _ in algorithm_configs}

                    for node_nf_count in NODE_NF_COUNTS:
                        if VERBOSE_PROGRESS:
                            print(f"      使用节点 NF 个数配置: NODE_DATA_{node_nf_count}")
                        for idx, (alg_name, alg_func) in enumerate(algorithm_configs, 1):
                            all_runs = []
                            for topo_instance_id, matrix in topologies.items():
                                if VERBOSE_PROGRESS:
                                    print(
                                        f"        拓扑实例 {topo_instance_id}/{len(topologies)} -> 算法 {idx}/4: {alg_name}",
                                        end=" ... ",
                                    )
                                for run_idx in range(NUM_RUNS):
                                    res = run_algorithm_on_matrix_silently(
                                        alg_func,
                                        alg_name,
                                        matrix,
                                        node_nf_count=node_nf_count,
                                        client_count=client_count,
                                        security_requirement=security_requirement,
                                    )
                                    # 标记拓扑信息，便于后续聚合/排查
                                    if isinstance(res, dict):
                                        res = dict(res)
                                        res["topology_size"] = node_size
                                        res["topology_degree"] = degree
                                        res["topology_instance_id"] = topo_instance_id
                                        res["security_requirement"] = security_requirement
                                    all_runs.append(res)
                                if VERBOSE_PROGRESS:
                                    print("✓")
                            per_client_results[alg_name][node_nf_count] = all_runs

                    per_degree_results[client_count] = per_client_results

                per_security_results[degree] = per_degree_results

            per_size_results[security_requirement] = per_security_results

        all_results[node_size] = per_size_results

    return all_results


def print_summary(all_results):
    """
    打印算法对比摘要：
    - 自变量仅为拓扑中的节点数量；
    - 对每个节点数量，在所有 security_requirement / degree / client 个数 (3/5/8/10) 和 NODE_DATA_2~5 组合上
      聚合，输出每个指标的最小值 / 最大值 / 平均值。
    
    指标：
      1. 总体成本 / 平均安全级别 (overall cost per security level)
      2. 平均安全级别 (average security)
      3. 总体成本 (安全开销 + 算力开销 + 带宽开销)
    """
    print("\n" + "=" * 80)
    print("按节点数量的算法对比摘要")
    print("=" * 80)

    algorithm_names = ALGORITHM_NAMES
    node_sizes = sorted(all_results.keys())

    metrics = METRICS

    for node_size in node_sizes:
        results_by_security = all_results[node_size]

        print(f"\n{'-' * 80}")
        print(
            f"{node_size} 节点拓扑，在安全需求 {SECURITY_REQUIREMENTS}、"
            f"度数 2/3/4/5、client 数量 {CLIENT_COUNTS} 与 NODE_DATA_2~5 上聚合"
        )
        print(f"{'-' * 80}")

        header = f"{'指标/统计':<20}"
        for alg_name in algorithm_names:
            header += f" {alg_name:<18}"
        print(header)
        print("-" * (20 + 19 * len(algorithm_names)))

        for metric_name, metric_id, fmt in metrics:
            for stat_name in ("最小值", "最大值", "平均值"):
                row = f"{metric_name}-{stat_name:<12}"
                for alg_name in algorithm_names:
                    values = []
                    # 遍历所有 security_requirement / degree / client_count / node_nf_count 组合
                    for security_results in results_by_security.values():
                        for degree_results in security_results.values():
                            for client_count, results_by_alg in degree_results.items():
                                nf_results = results_by_alg.get(alg_name, {})
                                for nf_count, run_list in nf_results.items():
                                    if not isinstance(run_list, list):
                                        continue
                                    for res in run_list:
                                        if not isinstance(res, dict):
                                            continue
                                        val = None
                                        try:
                                            if metric_id == "overall_cost_per_security":
                                                total_cost = res.get("total_cost")
                                                avg_sec = res.get("avg_node_security_level")
                                                if isinstance(total_cost, (int, float)) and isinstance(avg_sec, (int, float)) and avg_sec > 0:
                                                    val = total_cost / avg_sec
                                            elif metric_id == "avg_security":
                                                val = res.get("avg_node_security_level")
                                            elif metric_id == "overall_cost":
                                                val = res.get("total_cost")
                                            elif metric_id == "execution_time":
                                                val = res.get("execution_time")
                                        except Exception:  # noqa: BLE001
                                            val = None

                                    if isinstance(val, (int, float)):
                                        values.append(val)

                    if not values:
                        row += f" {'N/A':<18}"
                        continue

                    if stat_name == "最小值":
                        agg_val = min(values)
                    elif stat_name == "最大值":
                        agg_val = max(values)
                    else:
                        agg_val = sum(values) / len(values)

                    if fmt == "{}":
                        row += f" {fmt.format(int(agg_val)):<18}"
                    else:
                        row += f" {fmt.format(agg_val):<18}"
                print(row)

    print("=" * 80)


def _extract_metric_from_result(metric_id, res):
    """从单次算法结果字典中提取一个指标值。"""
    if not isinstance(res, dict):
        return None
    try:
        if metric_id == "overall_cost_per_security":
            total_cost = res.get("total_cost")
            avg_sec = res.get("avg_node_security_level")
            if isinstance(total_cost, (int, float)) and isinstance(avg_sec, (int, float)) and avg_sec > 0:
                return total_cost / avg_sec
        elif metric_id == "avg_security":
            return res.get("avg_node_security_level")
        elif metric_id == "overall_cost":
            return res.get("total_cost")
        elif metric_id == "execution_time":
            return res.get("execution_time")
    except Exception:  # noqa: BLE001
        return None
    return None


def aggregate_by_nodes(all_results):
    """
    结构：all_results[node_size][security_requirement][degree][client_count][alg_name][node_nf_count] = [res, ...]
    返回：{node_size: {alg_name: {metric_id: (min, max, avg)}}}
    """
    summary = {}

    for node_size, results_by_security in all_results.items():
        per_size = {}
        for alg_name in ALGORITHM_NAMES:
            per_size[alg_name] = {}

        for metric_name, metric_id, _ in METRICS:
            for alg_name in ALGORITHM_NAMES:
                values = []
                # 遍历所有 security_requirement / degree / client_count / node_nf_count / runs
                for security_results in results_by_security.values():
                    for degree_results in security_results.values():
                        for client_count, results_by_alg in degree_results.items():
                            nf_results = results_by_alg.get(alg_name, {})
                            for nf_count, run_list in nf_results.items():
                                if not isinstance(run_list, list):
                                    continue
                                for res in run_list:
                                    val = _extract_metric_from_result(metric_id, res)
                                    if isinstance(val, (int, float)):
                                        values.append(val)

                if not values:
                    per_size[alg_name][metric_id] = (None, None, None)
                else:
                    mn = min(values)
                    mx = max(values)
                    avg = sum(values) / len(values)
                    per_size[alg_name][metric_id] = (mn, mx, avg)

        summary[node_size] = per_size

    return summary


def write_nodes_sheet(wb, summary):
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet("按节点数量")

    # 表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    ws.append(["节点数量", "指标/统计"] + ALGORITHM_NAMES)
    for col in range(1, 2 + len(ALGORITHM_NAMES) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for node_size in sorted(summary.keys()):
        per_size = summary[node_size]
        # 每个 node_size 写一块：先一个空行标题
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(ALGORITHM_NAMES))
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"{node_size} 节点拓扑"
        cell.font = Font(bold=True)
        row_idx += 1

        for metric_name, metric_id, fmt in METRICS:
            for stat_name, stat_index in (("最小值", 0), ("最大值", 1), ("平均值", 2)):
                row = [node_size, f"{metric_name}-{stat_name}"]
                for alg_name in ALGORITHM_NAMES:
                    mn, mx, avg = per_size[alg_name].get(metric_id, (None, None, None))
                    val = (mn, mx, avg)[stat_index]
                    if val is None:
                        row.append("N/A")
                    else:
                        row.append(fmt.format(val))
                ws.append(row)
                row_idx += 1

    # 简单列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24


def main():
    print("=" * 80)
    print("按节点数量（20/30/40/50）对比四个算法")
    print("=" * 80)

    all_results = run_all_algorithms_by_nodes()
    print_summary(all_results)

    # 聚合并写入 Excel，命名规则：nodes + GENERATED_TOPOLOGIES_PER_COMBO + GENERATED_SEED_BASE + time
    print("\n开始聚合并写入 Excel ...")
    summary = aggregate_by_nodes(all_results)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    write_nodes_sheet(wb, summary)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_name = (
        f"nodes_{NUM_RUNS}_{GENERATED_TOPOLOGIES_PER_COMBO}_"
        f"{timestamp}.xlsx"
    )
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "results")
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, excel_name)
    wb.save(excel_path)

    print(f"✓ 已保存 Excel: {excel_path}")


if __name__ == "__main__":
    main()


